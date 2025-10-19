from fastapi.responses import FileResponse
from fastapi import APIRouter, Request, Query, Cookie, HTTPException
import sqlite3
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from typing import Optional
import csv
import io
from datetime import datetime

# ===== Ð˜ÐœÐŸÐžÐ Ð¢ Ð¦Ð•ÐÐ¢Ð ÐÐ›Ð˜Ð—ÐžÐ’ÐÐÐÐžÐ“Ðž Ð›ÐžÐ“Ð“Ð•Ð Ð =====
from logger import logger, log_info, log_error, log_warning

# ===== Ð˜ÐœÐŸÐžÐ Ð¢Ð« DATABASE =====
from database import (
    get_all_clients, get_or_create_client, get_all_bookings, get_all_messages, get_stats,
    get_analytics_data, get_funnel_data, update_booking_status,
    save_booking, get_user_by_session, get_chat_history, mark_messages_as_read,
    get_unread_messages_count, save_message, log_activity, update_client_status,
    get_client_by_id, update_client_info, pin_client, get_custom_statuses,
    create_custom_status, delete_custom_status,
    get_all_services
)

# ===== Ð˜ÐœÐŸÐžÐ Ð¢Ð« ÐšÐžÐÐ¤Ð˜Ð“Ð£Ð ÐÐ¦Ð˜Ð˜ =====
from config import SALON_INFO, CSS_VERSION, CLIENT_STATUSES, DATABASE_NAME

# ===== Ð˜ÐœÐŸÐžÐ Ð¢Ð« INSTAGRAM =====
from instagram import send_message


router = APIRouter()
templates = Jinja2Templates(directory="templates")


# ===== MIDDLEWARE Ð”Ð›Ð¯ ÐŸÐ ÐžÐ’Ð•Ð ÐšÐ˜ ÐÐ’Ð¢ÐžÐ Ð˜Ð—ÐÐ¦Ð˜Ð˜ =====
async def require_auth(session_token: Optional[str] = Cookie(None)):
    """ÐŸÑ€Ð¾Ð²ÐµÑ€Ð¸Ñ‚ÑŒ Ð°Ð²Ñ‚Ð¾Ñ€Ð¸Ð·Ð°Ñ†Ð¸ÑŽ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ"""
    if not session_token:
        return None
    return get_user_by_session(session_token)


# ===== Ð’Ð¡ÐŸÐžÐœÐžÐ“ÐÐ¢Ð•Ð›Ð¬ÐÐ«Ð• Ð¤Ð£ÐÐšÐ¦Ð˜Ð˜ =====
def get_total_unread():
    """ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ð¾Ð±Ñ‰ÐµÐµ ÐºÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾ Ð½ÐµÐ¿Ñ€Ð¾Ñ‡Ð¸Ñ‚Ð°Ð½Ð½Ñ‹Ñ… ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹"""
    clients = get_all_clients()
    total = 0
    for client in clients:
        total += get_unread_messages_count(client[0])
    return total


def get_all_statuses():
    """ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ð²ÑÐµ ÑÑ‚Ð°Ñ‚ÑƒÑÑ‹ (Ð±Ð°Ð·Ð¾Ð²Ñ‹Ðµ + ÐºÐ°ÑÑ‚Ð¾Ð¼Ð½Ñ‹Ðµ)"""
    statuses = CLIENT_STATUSES.copy()
    custom = get_custom_statuses()
    for status in custom:
        statuses[status[1]] = {
            "label": status[2],
            "color": status[3],
            "icon": status[4]
        }
    return statuses


def get_client_display_name(client):
    """ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ð¾Ñ‚Ð¾Ð±Ñ€Ð°Ð¶Ð°ÐµÐ¼Ð¾Ðµ Ð¸Ð¼Ñ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð° (Ð¿Ñ€Ð¸Ð¾Ñ€Ð¸Ñ‚ÐµÑ‚: Ð¸Ð¼Ñ > username > ID)"""
    if client[3]:  # name
        return client[3]
    elif client[1]:  # username
        return f"@{client[1]}"
    else:
        return client[0][:15] + "..."  # instagram_id


# ===== Ð”ÐÐ¨Ð‘ÐžÐ Ð” =====
@router.get("/admin", response_class=HTMLResponse)
async def admin_dashboard(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð“Ð»Ð°Ð²Ð½Ð°Ñ Ð¿Ð°Ð½ÐµÐ»ÑŒ ÑƒÐ¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¸Ñ CRM"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    stats = get_stats()
    analytics = get_analytics_data()
    funnel = get_funnel_data()
    recent_clients = get_all_clients()[:5]
    recent_bookings = get_all_bookings()[:5]
    unread_count = get_total_unread()

    return templates.TemplateResponse("admin/dashboard.html", {
        "request": request,
        "stats": stats,
        "analytics": analytics,
        "funnel": funnel,
        "recent_clients": recent_clients,
        "recent_bookings": recent_bookings,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION,
        "get_client_display_name": get_client_display_name,
    })


# ===== ÐšÐ›Ð˜Ð•ÐÐ¢Ð« =====
@router.get("/admin/clients", response_class=HTMLResponse)
async def admin_clients(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð£Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¸Ðµ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð°Ð¼Ð¸"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    clients = get_all_clients()
    stats = get_stats()
    unread_count = get_total_unread()
    all_statuses = get_all_statuses()

    return templates.TemplateResponse("admin/clients.html", {
        "request": request,
        "clients": clients,
        "stats": stats,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION,
        "client_statuses": all_statuses,
        "get_client_display_name": get_client_display_name
    })


@router.get("/admin/clients/{client_id}", response_class=HTMLResponse)
async def client_detail(request: Request, client_id: str, session_token: Optional[str] = Cookie(None)):
    """Ð”ÐµÑ‚Ð°Ð»ÑŒÐ½Ð°Ñ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ Ð¾ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ðµ"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    client = get_client_by_id(client_id)
    if not client:
        raise HTTPException(status_code=404, detail="ÐšÐ»Ð¸ÐµÐ½Ñ‚ Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½")

    history = get_chat_history(client_id, limit=50)
    bookings = [b for b in get_all_bookings() if b[1] == client_id]
    unread_count = get_total_unread()
    all_statuses = get_all_statuses()

    return templates.TemplateResponse("admin/client_detail.html", {
        "request": request,
        "client": client,
        "history": history,
        "bookings": bookings,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION,
        "client_statuses": all_statuses,
        "get_client_display_name": get_client_display_name
    })


# ===== Ð—ÐÐŸÐ˜Ð¡Ð˜ =====
@router.get("/admin/bookings", response_class=HTMLResponse)
async def admin_bookings(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð£Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¸Ðµ Ð·Ð°Ð¿Ð¸ÑÑÐ¼Ð¸"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    bookings = get_all_bookings()
    stats = get_stats()
    unread_count = get_total_unread()
    clients = get_all_clients()

    return templates.TemplateResponse("admin/bookings.html", {
        "request": request,
        "bookings": bookings,
        "clients": clients,
        "stats": stats,
        "get_client_display_name": get_client_display_name,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION
    })


@router.get("/admin/bookings/{booking_id}", response_class=HTMLResponse)
async def booking_detail(request: Request, booking_id: int, session_token: Optional[str] = Cookie(None)):
    """Ð”ÐµÑ‚Ð°Ð»ÑŒÐ½Ð°Ñ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ Ð¾ Ð·Ð°Ð¿Ð¸ÑÐ¸"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    conn = sqlite3.connect(DATABASE_NAME)
    c = conn.cursor()

    c.execute("""SELECT id, instagram_id, service_name, datetime, phone, name, status, created_at, revenue, notes
                 FROM bookings WHERE id = ?""", (booking_id,))
    booking = c.fetchone()

    if not booking:
        conn.close()
        raise HTTPException(status_code=404, detail="Ð—Ð°Ð¿Ð¸ÑÑŒ Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½Ð°")

    c.execute("""SELECT instagram_id, username, phone, name, first_contact, 
                 last_contact, total_messages, labels, status, lifetime_value,
                 profile_pic, notes, is_pinned 
                 FROM clients WHERE instagram_id = ?""", (booking[1],))
    client = c.fetchone()

    conn.close()

    unread_count = get_total_unread()

    return templates.TemplateResponse("admin/booking_detail.html", {
        "request": request,
        "booking": booking,
        "client": client,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION,
        "get_client_display_name": get_client_display_name
    })


# ===== Ð§ÐÐ¢ =====
@router.get("/admin/chat", response_class=HTMLResponse)
async def admin_chat(
    request: Request,
    client: Optional[str] = Query(None),
    session_token: Optional[str] = Cookie(None)
):
    """Ð˜Ð½Ñ‚ÐµÑ€Ñ„ÐµÐ¹Ñ Ñ‡Ð°Ñ‚Ð° Ñ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð°Ð¼Ð¸"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    clients = get_all_clients()
    all_statuses = get_all_statuses()

    chats = []
    for c in clients:
        history = get_chat_history(c[0], limit=1)
        unread = get_unread_messages_count(c[0])

        display_name = get_client_display_name(c)

        chats.append({
            "instagram_id": c[0],
            "username": c[1],
            "name": c[3],
            "display_name": display_name,
            "phone": c[2],
            "status": c[8],
            "profile_pic": c[10] if len(c) > 10 else None,
            "instagram_url": f"https://instagram.com/{c[1]}" if c[1] else None,
            "is_pinned": c[12] if len(c) > 12 else 0,
            "last_message": history[0][0] if history else "ÐÐµÑ‚ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹",
            "last_message_time": history[0][2][11:16] if history else "",
            "unread_count": unread
        })

    messages = []
    active_client_info = None

    if client:
        messages_raw = get_chat_history(client, limit=100)
        messages = [
            {
                "id": msg[4] if len(msg) > 4 else None,
                "message": msg[0],
                "sender": msg[1],
                "timestamp": msg[2],
                "type": msg[3] if len(msg) > 3 else "text"
            }
            for msg in messages_raw
        ]

        mark_messages_as_read(client, user["id"])

        client_data = next((c for c in clients if c[0] == client), None)
        if client_data:
            active_client_info = {
                "instagram_id": client_data[0],
                "username": client_data[1],
                "display_name": get_client_display_name(client_data),
                "phone": client_data[2],
                "name": client_data[3],
                "first_contact": client_data[4],
                "last_contact": client_data[5],
                "total_messages": client_data[6],
                "status": client_data[8],
                "lifetime_value": client_data[9],
                "profile_pic": client_data[10] if len(client_data) > 10 else None,
                "instagram_url": f"https://instagram.com/{client_data[1]}" if client_data[1] else None,
                "notes": client_data[11] if len(client_data) > 11 else ""
            }

    total_unread = sum(chat["unread_count"] for chat in chats)

    return templates.TemplateResponse("admin/chat.html", {
        "request": request,
        "chats": chats,
        "messages": messages,
        "active_client": client,
        "active_client_info": active_client_info,
        "unread_count": total_unread,
        "salon_info": SALON_INFO,
        "current_user": user,
        "css_version": CSS_VERSION,
        "client_statuses": all_statuses
    })


# ===== Ð¡ÐžÐžÐ‘Ð©Ð•ÐÐ˜Ð¯ =====
@router.get("/admin/messages", response_class=HTMLResponse)
async def admin_messages(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð˜ÑÑ‚Ð¾Ñ€Ð¸Ñ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    messages = get_all_messages(200)
    stats = get_stats()
    unread_count = get_total_unread()

    return templates.TemplateResponse("admin/messages.html", {
        "request": request,
        "messages": messages,
        "stats": stats,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION
    })


# ===== ÐÐÐÐ›Ð˜Ð¢Ð˜ÐšÐ =====
@router.get("/admin/analytics", response_class=HTMLResponse)
async def admin_analytics(request: Request, session_token: Optional[str] = Cookie(None)):
    """ÐÐ½Ð°Ð»Ð¸Ñ‚Ð¸ÐºÐ° Ð¸ Ð¾Ñ‚Ñ‡ÐµÑ‚Ñ‹"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    analytics = get_analytics_data()
    stats = get_stats()
    unread_count = get_total_unread()

    return templates.TemplateResponse("admin/analytics.html", {
        "request": request,
        "analytics": analytics,
        "stats": stats,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION
    })


# ===== Ð’ÐžÐ ÐžÐÐšÐ =====
@router.get("/admin/funnel", response_class=HTMLResponse)
async def admin_funnel(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð’Ð¾Ñ€Ð¾Ð½ÐºÐ° Ð¿Ñ€Ð¾Ð´Ð°Ð¶"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    funnel = get_funnel_data()
    stats = get_stats()
    unread_count = get_total_unread()

    return templates.TemplateResponse("admin/funnel.html", {
        "request": request,
        "funnel": funnel,
        "stats": stats,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION
    })


# ===== Ð£ÐŸÐ ÐÐ’Ð›Ð•ÐÐ˜Ð• Ð¡Ð¢ÐÐ¢Ð£Ð¡ÐÐœÐ˜ =====
@router.get("/admin/statuses", response_class=HTMLResponse)
async def manage_statuses(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð£Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¸Ðµ ÐºÐ°ÑÑ‚Ð¾Ð¼Ð½Ñ‹Ð¼Ð¸ ÑÑ‚Ð°Ñ‚ÑƒÑÐ°Ð¼Ð¸"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    all_statuses = get_all_statuses()
    custom_statuses = get_custom_statuses()
    unread_count = get_total_unread()

    return templates.TemplateResponse("admin/statuses.html", {
        "request": request,
        "all_statuses": all_statuses,
        "custom_statuses": custom_statuses,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION
    })


# ===== Ð£Ð¡Ð›Ð£Ð“Ð˜ =====
@router.get("/admin/services", response_class=HTMLResponse)
async def admin_services(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð£Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¸Ðµ ÑƒÑÐ»ÑƒÐ³Ð°Ð¼Ð¸"""
    user = await require_auth(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    services = get_all_services(active_only=False)
    stats = get_stats()
    unread_count = get_total_unread()

    return templates.TemplateResponse("admin/services.html", {
        "request": request,
        "services": services,
        "stats": stats,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION
    })


# ===== ÐŸÐžÐ›Ð¬Ð—ÐžÐ’ÐÐ¢Ð•Ð›Ð˜ =====
@router.get("/admin/users", response_class=HTMLResponse)
async def admin_users(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð£Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¸Ðµ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÑÐ¼Ð¸"""
    user = await require_auth(session_token)
    if not user or user["role"] != "admin":
        return RedirectResponse(url="/admin", status_code=302)
    
    from database import get_all_users
    
    users = get_all_users()
    unread_count = get_total_unread()
    
    return templates.TemplateResponse("admin/users.html", {
        "request": request,
        "users": users,
        "salon_info": SALON_INFO,
        "current_user": user,
        "unread_count": unread_count,
        "css_version": CSS_VERSION
    })


# ===== API Ð­ÐÐ”ÐŸÐžÐ˜ÐÐ¢Ð« =====
@router.get("/admin/api/stats")
async def get_stats_api(session_token: Optional[str] = Cookie(None)):
    """API Ð´Ð»Ñ Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ð¸Ñ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ¸"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return get_stats()


@router.get("/admin/api/unread-count")
async def get_unread_count_api(session_token: Optional[str] = Cookie(None)):
    """API Ð´Ð»Ñ Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ð¸Ñ ÐºÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð° Ð½ÐµÐ¿Ñ€Ð¾Ñ‡Ð¸Ñ‚Ð°Ð½Ð½Ñ‹Ñ… ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    return JSONResponse({"count": get_total_unread()})


@router.get("/admin/api/analytics")
async def get_analytics_api(
    period: int = Query(30),
    date_from: str = Query(None),
    date_to: str = Query(None),
    session_token: Optional[str] = Cookie(None)
):
    """API Ð´Ð»Ñ Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ð¸Ñ Ð°Ð½Ð°Ð»Ð¸Ñ‚Ð¸ÐºÐ¸ Ñ Ð¿ÐµÑ€Ð¸Ð¾Ð´Ð¾Ð¼"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    if date_from and date_to:
        return get_analytics_data(date_from=date_from, date_to=date_to)
    else:
        return get_analytics_data(days=period)


@router.get("/admin/api/funnel")
async def get_funnel_api(session_token: Optional[str] = Cookie(None)):
    """API Ð´Ð»Ñ Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ð¸Ñ Ð´Ð°Ð½Ð½Ñ‹Ñ… Ð²Ð¾Ñ€Ð¾Ð½ÐºÐ¸"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return get_funnel_data()


# ===== API Ð”Ð•Ð™Ð¡Ð¢Ð’Ð˜Ð¯ Ð¡ ÐšÐ›Ð˜Ð•ÐÐ¢ÐÐœÐ˜ =====
@router.post("/admin/api/clients/{client_id}/status")
async def update_client_status_api(
    client_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Ð˜Ð·Ð¼ÐµÐ½Ð¸Ñ‚ÑŒ ÑÑ‚Ð°Ñ‚ÑƒÑ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð°"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    new_status = data.get('status')

    if new_status:
        update_client_status(client_id, new_status)
        log_activity(user["id"], "update_client_status", "client", client_id,
                     f"Ð¡Ñ‚Ð°Ñ‚ÑƒÑ Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½ Ð½Ð° {new_status}")
        return JSONResponse({"success": True, "message": "Ð¡Ñ‚Ð°Ñ‚ÑƒÑ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð° Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½"})

    return JSONResponse({"success": False, "message": "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½ ÑÑ‚Ð°Ñ‚ÑƒÑ"}, status_code=400)


@router.post("/admin/api/clients/{client_id}/update")
async def update_client_api(
    client_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """ÐžÐ±Ð½Ð¾Ð²Ð¸Ñ‚ÑŒ Ð¸Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸ÑŽ Ð¾ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ðµ"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()

    success = update_client_info(
        client_id,
        name=data.get('name'),
        phone=data.get('phone'),
        notes=data.get('notes')
    )

    if success:
        log_activity(user["id"], "update_client_info",
                     "client", client_id, "Ð˜Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð°")
        return JSONResponse({"success": True, "message": "Ð˜Ð½Ñ„Ð¾Ñ€Ð¼Ð°Ñ†Ð¸Ñ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð°"})

    return JSONResponse({"success": False, "message": "ÐžÑˆÐ¸Ð±ÐºÐ° Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð¸Ñ"}, status_code=400)


@router.post("/admin/api/clients/{client_id}/notes")
async def save_client_notes(
    client_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Ð¡Ð¾Ñ…Ñ€Ð°Ð½Ð¸Ñ‚ÑŒ Ð·Ð°Ð¼ÐµÑ‚ÐºÐ¸ Ð¾ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ðµ"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    notes = data.get('notes')

    success = update_client_info(client_id, notes=notes)

    if success:
        log_activity(user["id"], "update_client_notes",
                     "client", client_id, "Ð—Ð°Ð¼ÐµÑ‚ÐºÐ¸ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ñ‹")
        return JSONResponse({"success": True, "message": "Ð—Ð°Ð¼ÐµÑ‚ÐºÐ¸ ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½Ñ‹"})

    return JSONResponse({"success": False, "message": "ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½Ð¸Ñ"}, status_code=400)


@router.post("/admin/api/clients/{client_id}/pin")
async def pin_client_api(
    client_id: str,
    session_token: Optional[str] = Cookie(None)
):
    """Ð—Ð°ÐºÑ€ÐµÐ¿Ð¸Ñ‚ÑŒ/Ð¾Ñ‚ÐºÑ€ÐµÐ¿Ð¸Ñ‚ÑŒ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð°"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    client = get_client_by_id(client_id)
    if not client:
        return JSONResponse({"success": False, "message": "ÐšÐ»Ð¸ÐµÐ½Ñ‚ Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½"}, status_code=404)

    is_pinned = client[12] if len(client) > 12 else 0
    pin_client(client_id, not is_pinned)

    log_activity(user["id"], "pin_client", "client", client_id,
                 f"{'Ð—Ð°ÐºÑ€ÐµÐ¿Ð»ÐµÐ½' if not is_pinned else 'ÐžÑ‚ÐºÑ€ÐµÐ¿Ð»ÐµÐ½'}")

    return JSONResponse({
        "success": True,
        "pinned": not is_pinned,
        "message": "Ð—Ð°ÐºÑ€ÐµÐ¿Ð»ÐµÐ½Ð¾" if not is_pinned else "ÐžÑ‚ÐºÑ€ÐµÐ¿Ð»ÐµÐ½Ð¾"
    })


# ===== API Ð”Ð•Ð™Ð¡Ð¢Ð’Ð˜Ð¯ Ð¡ Ð—ÐÐŸÐ˜Ð¡Ð¯ÐœÐ˜ =====
@router.post("/admin/api/bookings/create")
async def create_booking_api(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð½Ð¾Ð²ÑƒÑŽ Ð·Ð°Ð¿Ð¸ÑÑŒ Ð²Ñ€ÑƒÑ‡Ð½ÑƒÑŽ"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()

    try:
        instagram_id = data.get('instagram_id')
        service = data.get('service')
        datetime_str = f"{data.get('date')} {data.get('time')}"
        phone = data.get('phone')
        name = data.get('name')

        get_or_create_client(instagram_id, username=name)
        save_booking(instagram_id, service, datetime_str, phone, name)

        log_activity(user["id"], "create_booking", "booking", instagram_id,
                     f"Ð¡Ð¾Ð·Ð´Ð°Ð½Ð° Ð·Ð°Ð¿Ð¸ÑÑŒ: {service}")

        return JSONResponse({"success": True, "message": "Ð—Ð°Ð¿Ð¸ÑÑŒ ÑÐ¾Ð·Ð´Ð°Ð½Ð°"})
    except Exception as e:
        log_error(f"ÐžÑˆÐ¸Ð±ÐºÐ° ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ Ð·Ð°Ð¿Ð¸ÑÐ¸: {e}", "api", exc_info=True)
        return JSONResponse({"success": False, "message": str(e)}, status_code=400)


@router.post("/admin/api/bookings/{booking_id}/status")
async def update_booking_status_api(
    booking_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Ð˜Ð·Ð¼ÐµÐ½Ð¸Ñ‚ÑŒ ÑÑ‚Ð°Ñ‚ÑƒÑ Ð·Ð°Ð¿Ð¸ÑÐ¸"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    new_status = data.get('status')

    if new_status:
        success = update_booking_status(booking_id, new_status)
        if success:
            log_activity(user["id"], "update_booking_status", "booking", str(booking_id),
                         f"Ð¡Ñ‚Ð°Ñ‚ÑƒÑ Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½ Ð½Ð° {new_status}")
            return JSONResponse({"success": True, "message": "Ð¡Ñ‚Ð°Ñ‚ÑƒÑ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½"})
        else:
            return JSONResponse({"success": False, "message": "ÐžÑˆÐ¸Ð±ÐºÐ° Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð¸Ñ"}, status_code=400)

    return JSONResponse({"success": False, "message": "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½ ÑÑ‚Ð°Ñ‚ÑƒÑ"}, status_code=400)


@router.post("/admin/api/bookings/{booking_id}/notes")
async def save_booking_notes(
    booking_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Ð¡Ð¾Ñ…Ñ€Ð°Ð½Ð¸Ñ‚ÑŒ Ð·Ð°Ð¼ÐµÑ‚ÐºÐ¸ Ðº Ð·Ð°Ð¿Ð¸ÑÐ¸"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    notes = data.get('notes')

    conn = sqlite3.connect(DATABASE_NAME)
    c = conn.cursor()

    try:
        c.execute("UPDATE bookings SET notes = ? WHERE id = ?",
                  (notes, booking_id))
        conn.commit()
        conn.close()

        log_activity(user["id"], "update_booking_notes",
                     "booking", str(booking_id), "Ð—Ð°Ð¼ÐµÑ‚ÐºÐ¸ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ñ‹")
        return JSONResponse({"success": True, "message": "Ð—Ð°Ð¼ÐµÑ‚ÐºÐ¸ ÑÐ¾Ñ…Ñ€Ð°Ð½ÐµÐ½Ñ‹"})
    except Exception as e:
        conn.close()
        return JSONResponse({"success": False, "message": str(e)}, status_code=400)


from slowapi import Limiter
from slowapi.util import get_remote_address

limiter = Limiter(key_func=get_remote_address)

@router.post("/admin/api/chat/send")
@limiter.limit("30/minute")  # ÐœÐ°ÐºÑÐ¸Ð¼ÑƒÐ¼ 30 ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹ Ð² Ð¼Ð¸Ð½ÑƒÑ‚Ñƒ
async def send_chat_message_v2(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """ÐžÑ‚Ð¿Ñ€Ð°Ð²Ð¸Ñ‚ÑŒ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ Ñ rate limiting"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    instagram_id = data.get('instagram_id')
    message = data.get('message')

    if not instagram_id or not message:
        return JSONResponse({
            "success": False,
            "message": "ÐÐµÐ²ÐµÑ€Ð½Ñ‹Ðµ Ð´Ð°Ð½Ð½Ñ‹Ðµ"
        }, status_code=400)
    
    # ÐŸÑ€Ð¾Ð²ÐµÑ€ÐºÐ° Ð´Ð»Ð¸Ð½Ñ‹ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ñ
    if len(message) > 1000:
        return JSONResponse({
            "success": False,
            "message": "Ð¡Ð¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ ÑÐ»Ð¸ÑˆÐºÐ¾Ð¼ Ð´Ð»Ð¸Ð½Ð½Ð¾Ðµ (Ð¼Ð°ÐºÑ 1000 ÑÐ¸Ð¼Ð²Ð¾Ð»Ð¾Ð²)"
        }, status_code=400)

    result = await send_message(instagram_id, message)

    if "error" not in result:
        save_message(instagram_id, message, "bot")
        log_activity(user["id"], "send_message", "client",
                     instagram_id, f"ÐžÑ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¾: {message[:50]}")
        return JSONResponse({
            "success": True,
            "message": "Ð¡Ð¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¾"
        })
    else:
        return JSONResponse({
            "success": False,
            "message": f"ÐžÑˆÐ¸Ð±ÐºÐ° Instagram API: {result.get('error')}"
        }, status_code=500)

@router.get("/admin/api/chat/messages")
async def get_chat_messages(
    client: str = Query(...),
    session_token: Optional[str] = Cookie(None)
):
    """ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ñ Ñ‡Ð°Ñ‚Ð°"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    messages_raw = get_chat_history(client, limit=100)
    messages = [
        {
            "message": msg[0],
            "sender": msg[1],
            "timestamp": msg[2]
        }
        for msg in messages_raw
    ]

    mark_messages_as_read(client, user["id"])

    return JSONResponse({"messages": messages})


@router.delete("/admin/api/chat/message/{client_id}/{message_id}")
async def delete_message_v2(
    client_id: str,
    message_id: int,
    session_token: Optional[str] = Cookie(None)
):
    """Ð£Ð´Ð°Ð»Ð¸Ñ‚ÑŒ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ (Ð¢ÐžÐ›Ð¬ÐšÐž Ð¸Ð· CRM, Ð½Ðµ Ð¸Ð· Instagram)"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    conn = sqlite3.connect(DATABASE_NAME)
    c = conn.cursor()

    try:
        # ÐŸÑ€Ð¾Ð²ÐµÑ€ÑÐµÐ¼, Ñ‡Ñ‚Ð¾ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ ÑÑƒÑ‰ÐµÑÑ‚Ð²ÑƒÐµÑ‚
        c.execute("SELECT sender FROM chat_history WHERE instagram_id = ? AND id = ?",
                  (client_id, message_id))
        message = c.fetchone()
        
        if not message:
            conn.close()
            return JSONResponse({
                "success": False,
                "message": "Ð¡Ð¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½Ð¾"
            }, status_code=404)
        
        # Ð£Ð´Ð°Ð»ÑÐµÐ¼ Ð¸Ð· Ð‘Ð”
        c.execute("DELETE FROM chat_history WHERE instagram_id = ? AND id = ?",
                  (client_id, message_id))
        conn.commit()
        conn.close()

        log_activity(user["id"], "delete_message", "message", str(message_id), 
                     "Ð¡Ð¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ ÑƒÐ´Ð°Ð»ÐµÐ½Ð¾ Ð¸Ð· CRM (Ð¾ÑÑ‚Ð°Ð½ÐµÑ‚ÑÑ Ð² Instagram)")
        
        return JSONResponse({
            "success": True, 
            "message": "âš ï¸ Ð¡Ð¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ ÑƒÐ´Ð°Ð»ÐµÐ½Ð¾ Ð¸Ð· CRM. Ð’ÐÐ–ÐÐž: ÐžÐ½Ð¾ Ð¾ÑÑ‚Ð°Ð½ÐµÑ‚ÑÑ Ð² Instagram! Ð”Ð»Ñ Ð¿Ð¾Ð»Ð½Ð¾Ð³Ð¾ ÑƒÐ´Ð°Ð»ÐµÐ½Ð¸Ñ ÑƒÐ´Ð°Ð»Ð¸Ñ‚Ðµ Ð²Ñ€ÑƒÑ‡Ð½ÑƒÑŽ Ð² Instagram."
        })
    except Exception as e:
        conn.close()
        log_error(f"ÐžÑˆÐ¸Ð±ÐºÐ° ÑƒÐ´Ð°Ð»ÐµÐ½Ð¸Ñ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ñ: {e}", "chat", exc_info=True)
        return JSONResponse({"success": False, "message": str(e)}, status_code=400)


# ===== Ð—ÐÐ“Ð Ð£Ð—ÐšÐ Ð¤ÐÐ™Ð›ÐžÐ’ =====
@router.post("/admin/api/chat/upload")
async def upload_file_v2(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Ð—Ð°Ð³Ñ€ÑƒÐ·Ð¸Ñ‚ÑŒ Ñ„Ð°Ð¹Ð» Ñ Ð²Ð°Ð»Ð¸Ð´Ð°Ñ†Ð¸ÐµÐ¹"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    try:
        form = await request.form()
        file = form.get('file')
        instagram_id = form.get('instagram_id')
        is_image = form.get('is_image') == 'true'

        if not file or not instagram_id:
            return JSONResponse({
                "success": False,
                "message": "ÐÐµÐ´Ð¾ÑÑ‚Ð°Ñ‚Ð¾Ñ‡Ð½Ð¾ Ð´Ð°Ð½Ð½Ñ‹Ñ…"
            }, status_code=400)

        # Ð’ÐÐ›Ð˜Ð”ÐÐ¦Ð˜Ð¯ Ð ÐÐ—ÐœÐ•Ð Ð (Ð¼Ð°ÐºÑ 10MB)
        contents = await file.read()
        file_size_mb = len(contents) / (1024 * 1024)
        
        if file_size_mb > 10:
            return JSONResponse({
                "success": False,
                "message": f"Ð¤Ð°Ð¹Ð» ÑÐ»Ð¸ÑˆÐºÐ¾Ð¼ Ð±Ð¾Ð»ÑŒÑˆÐ¾Ð¹ ({file_size_mb:.1f}MB). ÐœÐ°ÐºÑÐ¸Ð¼ÑƒÐ¼ 10MB."
            }, status_code=400)

        filename = file.filename
        file_extension = filename.split('.')[-1].lower()

        # Ð’ÐÐ›Ð˜Ð”ÐÐ¦Ð˜Ð¯ Ð¢Ð˜ÐŸÐ Ð¤ÐÐ™Ð›Ð
        allowed_image_extensions = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
        allowed_file_extensions = {'pdf', 'doc', 'docx', 'txt', 'zip', 'rar', 'xlsx', 'xls'}
        
        if is_image:
            if file_extension not in allowed_image_extensions:
                return JSONResponse({
                    "success": False,
                    "message": f"ÐÐµÐ´Ð¾Ð¿ÑƒÑÑ‚Ð¸Ð¼Ñ‹Ð¹ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚ Ð¸Ð·Ð¾Ð±Ñ€Ð°Ð¶ÐµÐ½Ð¸Ñ. Ð Ð°Ð·Ñ€ÐµÑˆÐµÐ½Ð¾: {', '.join(allowed_image_extensions)}"
                }, status_code=400)
        else:
            if file_extension not in allowed_file_extensions:
                return JSONResponse({
                    "success": False,
                    "message": f"ÐÐµÐ´Ð¾Ð¿ÑƒÑÑ‚Ð¸Ð¼Ñ‹Ð¹ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚ Ñ„Ð°Ð¹Ð»Ð°. Ð Ð°Ð·Ñ€ÐµÑˆÐµÐ½Ð¾: {', '.join(allowed_file_extensions)}"
                }, status_code=400)

        # ÐŸÐ ÐžÐ’Ð•Ð ÐšÐ ÐÐ Ð’Ð˜Ð Ð£Ð¡Ð« (Ð±Ð°Ð·Ð¾Ð²Ð°Ñ - Ð¿Ð¾ magic bytes)
        if not is_safe_file(contents, file_extension):
            return JSONResponse({
                "success": False,
                "message": "Ð¤Ð°Ð¹Ð» Ð½Ðµ Ð¿Ñ€Ð¾ÑˆÑ‘Ð» Ð¿Ñ€Ð¾Ð²ÐµÑ€ÐºÑƒ Ð±ÐµÐ·Ð¾Ð¿Ð°ÑÐ½Ð¾ÑÑ‚Ð¸"
            }, status_code=400)

        import os
        upload_dir = "static/uploads/images" if is_image else "static/uploads/files"
        os.makedirs(upload_dir, exist_ok=True)

        # Ð‘Ð•Ð—ÐžÐŸÐÐ¡ÐÐžÐ• Ð˜ÐœÐ¯ Ð¤ÐÐ™Ð›Ð
        import re
        safe_filename = re.sub(r'[^a-zA-Z0-9._-]', '_', filename)
        
        timestamp = datetime.now().timestamp()
        unique_filename = f"{int(timestamp)}_{safe_filename}"
        file_path = os.path.join(upload_dir, unique_filename)

        with open(file_path, "wb") as f:
            f.write(contents)

        db_path = f"/static/uploads/{'images' if is_image else 'files'}/{unique_filename}"

        if is_image:
            message = f"ðŸ–¼ï¸ {unique_filename}"
            save_message(instagram_id, message, "bot", message_type="image")
        else:
            message = f"ðŸ“Ž {filename}"
            save_message(instagram_id, message, "bot", message_type="file")

        log_activity(user["id"], "send_file", "client",
                     instagram_id, f"{'Ð˜Ð·Ð¾Ð±Ñ€Ð°Ð¶ÐµÐ½Ð¸Ðµ' if is_image else 'Ð¤Ð°Ð¹Ð»'}: {filename}")

        return JSONResponse({
            "success": True,
            "message": "Ð¤Ð°Ð¹Ð» Ð·Ð°Ð³Ñ€ÑƒÐ¶ÐµÐ½",
            "filename": unique_filename,
            "file_path": db_path,
            "is_image": is_image
        })
    except Exception as e:
        log_error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐ¸ Ñ„Ð°Ð¹Ð»Ð°: {e}", "upload", exc_info=True)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


def is_safe_file(contents: bytes, extension: str) -> bool:
    """
    Ð‘Ð°Ð·Ð¾Ð²Ð°Ñ Ð¿Ñ€Ð¾Ð²ÐµÑ€ÐºÐ° Ñ„Ð°Ð¹Ð»Ð° Ð¿Ð¾ magic bytes
    
    Args:
        contents: ÑÐ¾Ð´ÐµÑ€Ð¶Ð¸Ð¼Ð¾Ðµ Ñ„Ð°Ð¹Ð»Ð°
        extension: Ñ€Ð°ÑÑˆÐ¸Ñ€ÐµÐ½Ð¸Ðµ Ñ„Ð°Ð¹Ð»Ð°
    
    Returns:
        True ÐµÑÐ»Ð¸ Ñ„Ð°Ð¹Ð» Ð±ÐµÐ·Ð¾Ð¿Ð°ÑÐµÐ½
    """
    # Magic bytes Ð´Ð»Ñ Ñ€Ð°Ð·Ð»Ð¸Ñ‡Ð½Ñ‹Ñ… Ñ‚Ð¸Ð¿Ð¾Ð² Ñ„Ð°Ð¹Ð»Ð¾Ð²
    magic_bytes = {
        'jpg': [b'\xff\xd8\xff'],
        'jpeg': [b'\xff\xd8\xff'],
        'png': [b'\x89\x50\x4e\x47'],
        'gif': [b'\x47\x49\x46\x38'],
        'webp': [b'\x52\x49\x46\x46'],
        'pdf': [b'\x25\x50\x44\x46'],
        'zip': [b'\x50\x4b\x03\x04', b'\x50\x4b\x05\x06'],
        'rar': [b'\x52\x61\x72\x21'],
        'doc': [b'\xd0\xcf\x11\xe0'],
        'docx': [b'\x50\x4b\x03\x04'],
        'xlsx': [b'\x50\x4b\x03\x04'],
        'xls': [b'\xd0\xcf\x11\xe0'],
    }
    
    if extension not in magic_bytes:
        # ÐÐµÐ¸Ð·Ð²ÐµÑÑ‚Ð½Ñ‹Ð¹ Ñ‚Ð¸Ð¿ - Ñ€Ð°Ð·Ñ€ÐµÑˆÐ°ÐµÐ¼ (Ð¼Ð¾Ð¶Ð½Ð¾ ÑƒÐ¶ÐµÑÑ‚Ð¾Ñ‡Ð¸Ñ‚ÑŒ)
        return True
    
    expected_bytes = magic_bytes[extension]
    
    for expected in expected_bytes:
        if contents.startswith(expected):
            return True
    
    # Magic bytes Ð½Ðµ ÑÐ¾Ð²Ð¿Ð°Ð´Ð°ÑŽÑ‚ - Ð¿Ð¾Ð´Ð¾Ð·Ñ€Ð¸Ñ‚ÐµÐ»ÑŒÐ½Ñ‹Ð¹ Ñ„Ð°Ð¹Ð»
    return False


@router.post("/admin/api/chat/voice")
async def upload_voice(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Ð—Ð°Ð³Ñ€ÑƒÐ·Ð¸Ñ‚ÑŒ Ð³Ð¾Ð»Ð¾ÑÐ¾Ð²Ð¾Ðµ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    try:
        form = await request.form()
        voice = form.get('voice')
        instagram_id = form.get('instagram_id')
        duration = form.get('duration', '0')

        if not voice or not instagram_id:
            return JSONResponse({"success": False, "message": "ÐÐµÐ´Ð¾ÑÑ‚Ð°Ñ‚Ð¾Ñ‡Ð½Ð¾ Ð´Ð°Ð½Ð½Ñ‹Ñ…"}, status_code=400)

        contents = await voice.read()

        import os
        upload_dir = "static/uploads/voice"
        os.makedirs(upload_dir, exist_ok=True)

        timestamp = datetime.now().timestamp()
        unique_filename = f"{int(timestamp)}_voice.webm"
        file_path = os.path.join(upload_dir, unique_filename)

        with open(file_path, "wb") as f:
            f.write(contents)

        db_path = f"/static/uploads/voice/{unique_filename}"

        message = f"ðŸŽ¤ {duration}s|{unique_filename}"
        save_message(instagram_id, message, "bot", message_type="voice")

        log_activity(user["id"], "send_voice", "client",
                     instagram_id, f"Ð“Ð¾Ð»Ð¾ÑÐ¾Ð²Ð¾Ðµ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ðµ ({duration}s)")

        return JSONResponse({
            "success": True,
            "message": "Ð“Ð¾Ð»Ð¾ÑÐ¾Ð²Ð¾Ðµ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¾",
            "filename": unique_filename,
            "file_path": db_path,
            "duration": duration
        })
    except Exception as e:
        log_error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð·Ð°Ð³Ñ€ÑƒÐ·ÐºÐ¸ Ð³Ð¾Ð»Ð¾ÑÐ¾Ð²Ð¾Ð³Ð¾: {e}", "upload", exc_info=True)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


@router.get("/admin/api/uploads/{file_type}/{filename}")
async def get_uploaded_file(
    file_type: str,
    filename: str,
    session_token: Optional[str] = Cookie(None)
):
    """ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ð·Ð°Ð³Ñ€ÑƒÐ¶ÐµÐ½Ð½Ñ‹Ð¹ Ñ„Ð°Ð¹Ð»"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    import os
    
    file_path = os.path.join("static", "uploads", file_type, filename)
    
    if not os.path.exists(file_path):
        log_error(f"Ð¤Ð°Ð¹Ð» Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½: {file_path}", "upload")
        raise HTTPException(status_code=404, detail=f"Ð¤Ð°Ð¹Ð» Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½: {filename}")
    
    log_info(f"ÐžÑ‚Ð´Ð°Ñ‘Ð¼ Ñ„Ð°Ð¹Ð»: {file_path}", "upload")

    ext = filename.split('.')[-1].lower()
    mime_mapping = {
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'gif': 'image/gif',
        'webp': 'image/webp',
        'webm': 'audio/webm',
        'mp3': 'audio/mpeg',
        'pdf': 'application/pdf',
        'doc': 'application/msword',
        'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    }
    
    media_type = mime_mapping.get(ext, 'application/octet-stream')

    return FileResponse(
        file_path,
        media_type=media_type,
        filename=filename,
        headers={
            "Cache-Control": "public, max-age=86400",
        }
    )


@router.get("/admin/api/chats-update")
async def get_chats_update(session_token: Optional[str] = Cookie(None)):
    """API Ð´Ð»Ñ Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½Ð¸Ñ Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð¸Ð¹ Ñ‡Ð°Ñ‚Ð¾Ð²"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    clients = get_all_clients()
    chats = []
    total_unread = 0

    for c in clients:
        history = get_chat_history(c[0], limit=1)
        unread = get_unread_messages_count(c[0])
        total_unread += unread

        chats.append({
            "instagram_id": c[0],
            "username": c[1],
            "name": c[3],
            "display_name": get_client_display_name(c),
            "last_message": history[0][0] if history else "",
            "last_message_time": history[0][2][11:16] if history else "",
            "unread_count": unread,
            "has_new_message": unread > 0
        })

    return JSONResponse({
        "chats": chats,
        "total_unread": total_unread
    })


# ===== API Ð¡Ð¢ÐÐ¢Ð£Ð¡Ð« =====
@router.post("/admin/api/statuses/create")
async def create_status_api(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð½Ð¾Ð²Ñ‹Ð¹ ÑÑ‚Ð°Ñ‚ÑƒÑ"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    success = create_custom_status(
        data['status_key'],
        data['status_label'],
        data['status_color'],
        data['status_icon'],
        user['id']
    )

    if success:
        log_activity(user["id"], "create_status", "status",
                     data['status_key'], f"Ð¡Ð¾Ð·Ð´Ð°Ð½ ÑÑ‚Ð°Ñ‚ÑƒÑ {data['status_label']}")
        return JSONResponse({"success": True, "message": "Ð¡Ñ‚Ð°Ñ‚ÑƒÑ ÑÐ¾Ð·Ð´Ð°Ð½"})

    return JSONResponse({"success": False, "message": "Ð¡Ñ‚Ð°Ñ‚ÑƒÑ Ñ Ñ‚Ð°ÐºÐ¸Ð¼ ÐºÐ»ÑŽÑ‡Ð¾Ð¼ ÑƒÐ¶Ðµ ÑÑƒÑ‰ÐµÑÑ‚Ð²ÑƒÐµÑ‚"}, status_code=400)


@router.post("/admin/api/statuses/{status_key}/delete")
async def delete_status_api(status_key: str, session_token: Optional[str] = Cookie(None)):
    """Ð£Ð´Ð°Ð»Ð¸Ñ‚ÑŒ ÑÑ‚Ð°Ñ‚ÑƒÑ"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    success = delete_custom_status(status_key)

    if success:
        log_activity(user["id"], "delete_status",
                     "status", status_key, "Ð¡Ñ‚Ð°Ñ‚ÑƒÑ ÑƒÐ´Ð°Ð»ÐµÐ½")
        return JSONResponse({"success": True, "message": "Ð¡Ñ‚Ð°Ñ‚ÑƒÑ ÑƒÐ´Ð°Ð»ÐµÐ½"})

    return JSONResponse({"success": False, "message": "ÐžÑˆÐ¸Ð±ÐºÐ° ÑƒÐ´Ð°Ð»ÐµÐ½Ð¸Ñ"}, status_code=400)


# ===== API Ð£Ð¡Ð›Ð£Ð“Ð˜ =====
@router.post("/admin/api/services/create")
async def create_service_api(request: Request, session_token: Optional[str] = Cookie(None)):
    """Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð½Ð¾Ð²ÑƒÑŽ ÑƒÑÐ»ÑƒÐ³Ñƒ"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    from database import create_service
    
    data = await request.json()
    
    benefits = data.get('benefits', '').split('|') if data.get('benefits') else []
    
    success = create_service(
        service_key=data['service_key'],
        name=data['name'],
        name_ru=data.get('name_ru'),
        price=float(data['price']),
        currency=data.get('currency', 'AED'),
        category=data['category'],
        description=data.get('description'),
        description_ru=data.get('description_ru'),
        benefits=benefits
    )

    if success:
        log_activity(user["id"], "create_service", "service", data['service_key'], 
                     f"Ð¡Ð¾Ð·Ð´Ð°Ð½Ð° ÑƒÑÐ»ÑƒÐ³Ð° {data['name']}")
        return JSONResponse({"success": True, "message": "Ð£ÑÐ»ÑƒÐ³Ð° ÑÐ¾Ð·Ð´Ð°Ð½Ð°"})

    return JSONResponse({"success": False, "message": "Ð£ÑÐ»ÑƒÐ³Ð° Ñ Ñ‚Ð°ÐºÐ¸Ð¼ ÐºÐ»ÑŽÑ‡Ð¾Ð¼ ÑƒÐ¶Ðµ ÑÑƒÑ‰ÐµÑÑ‚Ð²ÑƒÐµÑ‚"}, status_code=400)


@router.post("/admin/api/services/{service_id}/update")
async def update_service_api(service_id: int, request: Request, session_token: Optional[str] = Cookie(None)):
    """ÐžÐ±Ð½Ð¾Ð²Ð¸Ñ‚ÑŒ ÑƒÑÐ»ÑƒÐ³Ñƒ"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    from database import update_service
    
    data = await request.json()
    
    if 'benefits' in data and isinstance(data['benefits'], str):
        data['benefits'] = data['benefits'].split('|') if data['benefits'] else []
    
    success = update_service(service_id, **data)

    if success:
        log_activity(user["id"], "update_service", "service", str(service_id), 
                     "Ð£ÑÐ»ÑƒÐ³Ð° Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð°")
        return JSONResponse({"success": True, "message": "Ð£ÑÐ»ÑƒÐ³Ð° Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð°"})

    return JSONResponse({"success": False, "message": "ÐžÑˆÐ¸Ð±ÐºÐ° Ð¾Ð±Ð½Ð¾Ð²Ð»ÐµÐ½Ð¸Ñ"}, status_code=400)


@router.post("/admin/api/services/{service_id}/delete")
async def delete_service_api(service_id: int, session_token: Optional[str] = Cookie(None)):
    """Ð£Ð´Ð°Ð»Ð¸Ñ‚ÑŒ ÑƒÑÐ»ÑƒÐ³Ñƒ"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    from database import delete_service
    
    success = delete_service(service_id)

    if success:
        log_activity(user["id"], "delete_service", "service", str(service_id), 
                     "Ð£ÑÐ»ÑƒÐ³Ð° ÑƒÐ´Ð°Ð»ÐµÐ½Ð°")
        return JSONResponse({"success": True, "message": "Ð£ÑÐ»ÑƒÐ³Ð° ÑƒÐ´Ð°Ð»ÐµÐ½Ð°"})

    return JSONResponse({"success": False, "message": "ÐžÑˆÐ¸Ð±ÐºÐ° ÑƒÐ´Ð°Ð»ÐµÐ½Ð¸Ñ"}, status_code=400)


@router.get("/admin/api/services/{service_id}")
async def get_service_api(service_id: int, session_token: Optional[str] = Cookie(None)):
    """ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÑƒÑÐ»ÑƒÐ³Ñƒ Ð¿Ð¾ ID"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    conn = sqlite3.connect(DATABASE_NAME)
    c = conn.cursor()
    
    c.execute("SELECT * FROM services WHERE id = ?", (service_id,))
    service = c.fetchone()
    conn.close()
    
    if not service:
        return JSONResponse({"error": "Service not found"}, status_code=404)
    
    return JSONResponse({
        "id": service[0],
        "service_key": service[1],
        "name": service[2],
        "name_ru": service[3],
        "name_ar": service[4],
        "price": service[5],
        "currency": service[6],
        "category": service[7],
        "description": service[8],
        "description_ru": service[9],
        "description_ar": service[10],
        "benefits": service[11],
        "is_active": service[12]
    })


# ===== API ÐŸÐžÐ›Ð¬Ð—ÐžÐ’ÐÐ¢Ð•Ð›Ð˜ =====
@router.post("/admin/api/users/{user_id}/delete")
async def delete_user_api(user_id: int, session_token: Optional[str] = Cookie(None)):
    """Ð£Ð´Ð°Ð»Ð¸Ñ‚ÑŒ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ"""
    user = await require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    if user["id"] == user_id:
        return JSONResponse({"success": False, "message": "ÐÐµÐ»ÑŒÐ·Ñ ÑƒÐ´Ð°Ð»Ð¸Ñ‚ÑŒ ÑÐµÐ±Ñ"}, status_code=400)
    
    from database import delete_user
    
    success = delete_user(user_id)
    
    if success:
        log_activity(user["id"], "delete_user", "user", str(user_id), "ÐŸÐ¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÑŒ ÑƒÐ´Ð°Ð»ÐµÐ½")
        return JSONResponse({"success": True, "message": "ÐŸÐ¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»ÑŒ ÑƒÐ´Ð°Ð»ÐµÐ½"})
    
    return JSONResponse({"success": False, "message": "ÐžÑˆÐ¸Ð±ÐºÐ° ÑƒÐ´Ð°Ð»ÐµÐ½Ð¸Ñ"}, status_code=400)


# ===== Ð­ÐšÐ¡ÐŸÐžÐ Ð¢ Ð”ÐÐÐÐ«Ð¥ =====
@router.get("/admin/api/export/clients")
async def export_clients(
    format: str = Query("csv"),
    session_token: Optional[str] = Cookie(None)
):
    """Ð­ÐºÑÐ¿Ð¾Ñ€Ñ‚ ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð² Ð² CSV/Excel/PDF"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    clients = get_all_clients()

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(['ID', 'Ð˜Ð¼Ñ', 'Username', 'Ð¢ÐµÐ»ÐµÑ„Ð¾Ð½', 'Ð¡Ñ‚Ð°Ñ‚ÑƒÑ', 'Ð¡Ð¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹',
                        'ÐŸÐµÑ€Ð²Ñ‹Ð¹ ÐºÐ¾Ð½Ñ‚Ð°ÐºÑ‚', 'ÐŸÐ¾ÑÐ»ÐµÐ´Ð½Ð¸Ð¹ ÐºÐ¾Ð½Ñ‚Ð°ÐºÑ‚', 'LTV'])

        for client in clients:
            writer.writerow([
                client[0], client[3] or '', client[1] or '', client[2] or '',
                client[8], client[6], client[4], client[5], client[9]
            ])

        output.seek(0)
        csv_content = '\ufeff' + output.getvalue()

        return StreamingResponse(
            iter([csv_content.encode('utf-8')]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename=clients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                "Content-Type": "text/csv; charset=utf-8"
            }
        )

    elif format == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ÐšÐ»Ð¸ÐµÐ½Ñ‚Ñ‹"

            headers = ['ID', 'Ð˜Ð¼Ñ', 'Username', 'Ð¢ÐµÐ»ÐµÑ„Ð¾Ð½', 'Ð¡Ñ‚Ð°Ñ‚ÑƒÑ', 'Ð¡Ð¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ð¹',
                       'ÐŸÐµÑ€Ð²Ñ‹Ð¹ ÐºÐ¾Ð½Ñ‚Ð°ÐºÑ‚', 'ÐŸÐ¾ÑÐ»ÐµÐ´Ð½Ð¸Ð¹ ÐºÐ¾Ð½Ñ‚Ð°ÐºÑ‚', 'LTV']
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="667EEA", end_color="667EEA", fill_type="solid")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")

            for client in clients:
                ws.append([
                    client[0], client[3] or '', client[1] or '', client[2] or '',
                    client[8], client[6], client[4], client[5], client[9]
                ])

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=clients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                }
            )

        except ImportError:
            return JSONResponse({"error": "openpyxl not installed. Run: pip install openpyxl"}, status_code=500)

    elif format == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib.enums import TA_CENTER

            output = io.BytesIO()
            doc = SimpleDocTemplate(
                output,
                pagesize=A4,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )

            elements = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#667eea'),
                spaceAfter=10,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.HexColor('#6b7280'),
                spaceAfter=20,
                alignment=TA_CENTER
            )

            elements.append(Paragraph("M.Le Diamant Beauty Lounge", title_style))
            elements.append(Paragraph("CLIENT DATABASE REPORT", subtitle_style))
            elements.append(Paragraph(
                f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}",
                subtitle_style
            ))
            elements.append(Spacer(1, 0.3*cm))

            stats_data = [
                ['Total Clients', 'Active', 'LTV Total', 'Avg Messages'],
                [
                    str(len(clients)),
                    str(len([c for c in clients if c[8] in ['new', 'lead', 'customer']])),
                    f"{sum([c[9] for c in clients if c[9]])} AED",
                    f"{sum([c[6] for c in clients]) // len(clients) if clients else 0}"
                ]
            ]

            stats_table = Table(stats_data, colWidths=[3*cm, 3*cm, 3*cm, 3*cm])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f3f4f6')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ]))
            elements.append(stats_table)
            elements.append(Spacer(1, 0.4*cm))

            data = [['Name', 'Username', 'Phone', 'Status', 'Messages', 'LTV']]

            for client in clients:
                name = client[3] if client[3] else (client[1] if client[1] else client[0][:15])
                username = f"@{client[1]}" if client[1] else "-"
                phone = client[2] if client[2] else "-"
                status = client[8].upper() if client[8] else "NEW"
                messages = str(client[6])
                ltv = f"{client[9]} AED" if client[9] else "0 AED"

                data.append([name[:20], username[:15], phone[:15], status, messages, ltv])

            col_widths = [4*cm, 3*cm, 3*cm, 2.5*cm, 2*cm, 2.5*cm]
            client_table = Table(data, colWidths=col_widths, repeatRows=1)

            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('ALIGN', (4, 1), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.white, colors.HexColor('#f9fafb')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#667eea')),
            ])

            client_table.setStyle(table_style)
            elements.append(client_table)

            doc.build(elements)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename=clients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                }
            )
        except ImportError:
            return JSONResponse({"error": "reportlab not installed. Run: pip install reportlab"}, status_code=500)
    
    return JSONResponse({"error": "Format not supported"}, status_code=400)


@router.get("/admin/api/export/analytics")
async def export_analytics(
    format: str = Query("csv"),
    period: int = Query(30),
    date_from: str = Query(None),
    date_to: str = Query(None),
    session_token: Optional[str] = Cookie(None)
):
    """Ð­ÐºÑÐ¿Ð¾Ñ€Ñ‚ Ð°Ð½Ð°Ð»Ð¸Ñ‚Ð¸ÐºÐ¸"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    if date_from and date_to:
        analytics = get_analytics_data(date_from=date_from, date_to=date_to)
    else:
        analytics = get_analytics_data(days=period)

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(['ÐžÑ‚Ñ‡ÐµÑ‚ Ð¿Ð¾ Ð°Ð½Ð°Ð»Ð¸Ñ‚Ð¸ÐºÐµ'])
        writer.writerow([])
        writer.writerow(['Ð—Ð°Ð¿Ð¸ÑÐ¸ Ð¿Ð¾ Ð´Ð½ÑÐ¼'])
        writer.writerow(['Ð”Ð°Ñ‚Ð°', 'ÐšÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾'])
        for row in analytics['bookings_by_day']:
            writer.writerow(row)

        writer.writerow([])
        writer.writerow(['Ð¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ° Ð¿Ð¾ ÑƒÑÐ»ÑƒÐ³Ð°Ð¼'])
        writer.writerow(['Ð£ÑÐ»ÑƒÐ³Ð°', 'ÐšÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾', 'Ð”Ð¾Ñ…Ð¾Ð´'])
        for row in analytics['services_stats']:
            writer.writerow(row)

        writer.writerow([])
        writer.writerow(['Ð¡Ñ‚Ð°Ñ‚ÑƒÑÑ‹ Ð·Ð°Ð¿Ð¸ÑÐµÐ¹'])
        writer.writerow(['Ð¡Ñ‚Ð°Ñ‚ÑƒÑ', 'ÐšÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾'])
        for row in analytics['status_stats']:
            writer.writerow(row)

        writer.writerow([])
        writer.writerow(['Ð¡Ñ€ÐµÐ´Ð½ÐµÐµ Ð²Ñ€ÐµÐ¼Ñ Ð¾Ñ‚Ð²ÐµÑ‚Ð° (Ð¼Ð¸Ð½)',
                        analytics['avg_response_time']])

        output.seek(0)
        csv_content = '\ufeff' + output.getvalue()

        return StreamingResponse(
            iter([csv_content.encode('utf-8')]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename=analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                "Content-Type": "text/csv; charset=utf-8"
            }
        )

    return JSONResponse({"error": "Format not supported"}, status_code=400)


@router.get("/admin/api/export/bookings")
async def export_bookings(
    format: str = Query("csv"),
    session_token: Optional[str] = Cookie(None)
):
    """Ð­ÐºÑÐ¿Ð¾Ñ€Ñ‚ Ð·Ð°Ð¿Ð¸ÑÐµÐ¹ Ð² CSV/Excel"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    bookings = get_all_bookings()

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['ID', 'ÐšÐ»Ð¸ÐµÐ½Ñ‚ ID', 'Ð£ÑÐ»ÑƒÐ³Ð°', 'Ð”Ð°Ñ‚Ð°/Ð’Ñ€ÐµÐ¼Ñ',
                        'Ð¢ÐµÐ»ÐµÑ„Ð¾Ð½', 'Ð˜Ð¼Ñ', 'Ð¡Ñ‚Ð°Ñ‚ÑƒÑ', 'Ð¡Ð¾Ð·Ð´Ð°Ð½Ð¾', 'Ð”Ð¾Ñ…Ð¾Ð´'])

        for booking in bookings:
            writer.writerow([
                booking[0], booking[1], booking[2], booking[3],
                booking[4], booking[5], booking[6], booking[7], booking[8] if len(
                    booking) > 8 else 0
            ])

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=bookings_{datetime.now().strftime('%Y%m%d')}.csv"}
        )

    return JSONResponse({"error": "Format not supported"}, status_code=400)


@router.get("/admin/api/chat/messages")
async def get_chat_messages_paginated(
    client: str = Query(...),
    limit: int = Query(50, le=100),
    offset: int = Query(0),
    session_token: Optional[str] = Cookie(None)
):
    """ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ñ Ñ Ð¿Ð°Ð³Ð¸Ð½Ð°Ñ†Ð¸ÐµÐ¹"""
    user = await require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    from database import DATABASE_NAME
    
    conn = sqlite3.connect(DATABASE_NAME)
    c = conn.cursor()
    
    # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ Ð¾Ð±Ñ‰ÐµÐµ ÐºÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾
    c.execute("""SELECT COUNT(*) FROM chat_history 
                 WHERE instagram_id = ?""", (client,))
    total = c.fetchone()[0]
    
    # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ñ Ñ Ð¿Ð°Ð³Ð¸Ð½Ð°Ñ†Ð¸ÐµÐ¹
    c.execute("""SELECT message, sender, timestamp, message_type, id
                 FROM chat_history 
                 WHERE instagram_id = ? 
                 ORDER BY timestamp DESC 
                 LIMIT ? OFFSET ?""",
              (client, limit, offset))
    
    messages_raw = c.fetchall()
    conn.close()
    
    messages = [
        {
            "id": msg[4],
            "message": msg[0],
            "sender": msg[1],
            "timestamp": msg[2],
            "type": msg[3]
        }
        for msg in reversed(messages_raw)
    ]

    mark_messages_as_read(client, user["id"])

    return JSONResponse({
        "messages": messages,
        "total": total,
        "has_more": (offset + limit) < total
    })